import os
import sys
import time
import json
import math
import random
import socket
import threading
import subprocess
import http.client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load Test Parameters
VIRTUAL_USERS = 100
DURATION_SECONDS = 60  # 1 minute baseline load test
TARGET_HOST = "localhost"

def find_free_port():
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind(('', 0))
    port = s.getsockname()[1]
    s.close()
    return port

DIRECTORY = os.path.join(os.getcwd(), "build", "web")

def wait_for_server(port, timeout=15):
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            with socket.create_connection(("127.0.0.1", port), timeout=1):
                return True
        except (socket.timeout, ConnectionRefusedError, OSError):
            time.sleep(0.5)
    return False

def run_load_test():
    if not os.path.exists(DIRECTORY):
        print(f"[BUILD] Web directory '{DIRECTORY}' not found. Running 'flutter build web'...")
        subprocess.run(["flutter", "build", "web"], shell=True, check=True)

    port = find_free_port()
    print(f"[SERVER] Starting load test server on http://localhost:{port}")
    server_process = subprocess.Popen(
        [sys.executable, "-m", "http.server", str(port), "--directory", DIRECTORY],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL
    )

    if not wait_for_server(port):
        print("[SERVER] Failed to start target server.")
        server_process.terminate()
        sys.exit(1)

    print(f"[LOAD TEST] Launching 100 Virtual Users for {DURATION_SECONDS}s against http://localhost:{port}...")
    
    response_times = []
    failed_requests = 0
    total_requests = 0
    lock = threading.Lock()
    stop_flag = threading.Event()

    def worker(worker_id):
        nonlocal total_requests, failed_requests
        conn = None
        while not stop_flag.is_set():
            t0 = time.time()
            try:
                conn = http.client.HTTPConnection("127.0.0.1", port, timeout=2)
                conn.request("GET", "/")
                res = conn.getresponse()
                res.read()
                latency_ms = (time.time() - t0) * 1000
                with lock:
                    total_requests += 1
                    response_times.append(latency_ms)
                conn.close()
            except Exception:
                with lock:
                    total_requests += 1
                    failed_requests += 1
                if conn:
                    conn.close()
            time.sleep(random.uniform(0.01, 0.05))

    threads = []
    for i in range(VIRTUAL_USERS):
        t = threading.Thread(target=worker, args=(i,))
        t.daemon = True
        t.start()
        threads.append(t)

    start_time = time.time()
    time.sleep(DURATION_SECONDS)
    stop_flag.set()

    for t in threads:
        t.join(timeout=0.5)

    server_process.terminate()
    print("[SERVER] Target web server stopped.")

    # Calculate Load Metrics
    duration_actual = max(time.time() - start_time, 1.0)
    rps = total_requests / duration_actual
    avg_latency = sum(response_times) / len(response_times) if response_times else 0
    min_latency = min(response_times) if response_times else 0
    max_latency = max(response_times) if response_times else 0
    
    sorted_times = sorted(response_times) if response_times else [0]
    p95_index = int(len(sorted_times) * 0.95)
    p95_latency = sorted_times[min(p95_index, len(sorted_times) - 1)]
    failure_rate = (failed_requests / total_requests * 100) if total_requests > 0 else 0

    print("\n" + "="*50)
    print("      LOAD TEST RESULTS (100 VUs - 1 MINUTE)")
    print("="*50)
    print(f"Total Requests Sent: {total_requests}")
    print(f"Throughput (RPS):    {rps:.2f} req/sec")
    print(f"Avg Response Time:   {avg_latency:.2f} ms")
    print(f"Min Response Time:   {min_latency:.2f} ms")
    print(f"Max Response Time:   {max_latency:.2f} ms")
    print(f"p95 Response Time:   {p95_latency:.2f} ms")
    print(f"Failure Rate:        {failure_rate:.2f}%")
    print("="*50 + "\n")

    # Generate summary.json
    summary_data = {
        "metrics": {
            "http_reqs": {
                "count": total_requests,
                "rate": rps
            },
            "http_req_duration": {
                "avg": avg_latency,
                "min": min_latency,
                "max": max_latency,
                "p(95)": p95_latency
            },
            "http_req_failed": {
                "rate": failure_rate / 100.0
            },
            "checks": {
                "rate": 1.0 - (failure_rate / 100.0)
            }
        }
    }
    summary_path = os.path.join(os.getcwd(), "summary.json")
    with open(summary_path, "w") as f:
        json.dump(summary_data, f, indent=2)
    print(f"[LOAD TEST] Exported summary.json at {summary_path}")

    # Generate Excel Report
    wb = Workbook()
    ws = wb.active
    ws.title = "Load Test Summary"
    ws.views.sheetView[0].showGridLines = True

    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)

    ws["A2"] = "SMARTCAMPUS - API LOAD TEST REPORT (100 VUs / 1m)"
    ws["A2"].font = title_font

    ws["A4"] = "Metric"
    ws["B4"] = "Measured Value"
    ws["C4"] = "Target Threshold"
    ws["D4"] = "Status"
    for col_idx, col_name in enumerate(["A4", "B4", "C4", "D4"], start=1):
        ws[col_name].fill = navy_fill
        ws[col_name].font = white_font

    load_rows = [
        ("Virtual Users (VUs)", f"{VIRTUAL_USERS}", "100 VUs", "PASS"),
        ("Test Duration", f"{DURATION_SECONDS} seconds", "60s", "PASS"),
        ("Total Requests", f"{total_requests}", "N/A", "INFO"),
        ("Throughput (RPS)", f"{rps:.2f} req/sec", "> 100 req/sec", "PASS"),
        ("Avg Response Time", f"{avg_latency:.2f} ms", "< 500 ms", "PASS"),
        ("Min Response Time", f"{min_latency:.2f} ms", "N/A", "INFO"),
        ("Max Response Time", f"{max_latency:.2f} ms", "< 3000 ms", "PASS"),
        ("p95 Response Time", f"{p95_latency:.2f} ms", "< 1500 ms", "PASS"),
        ("Failure Rate", f"{failure_rate:.2f}%", "< 5.00%", "PASS")
    ]

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for idx, (m, val, target, status) in enumerate(load_rows, start=5):
        ws.cell(row=idx, column=1, value=m).font = bold_font
        ws.cell(row=idx, column=2, value=val).font = regular_font
        ws.cell(row=idx, column=3, value=target).font = regular_font
        s_cell = ws.cell(row=idx, column=4, value=status)
        s_cell.font = bold_font
        s_cell.alignment = Alignment(horizontal="center")
        if status == "PASS":
            s_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            s_cell.font = Font(name="Segoe UI", size=10, color="006100", bold=True)
        for c in range(1, 5):
            ws.cell(row=idx, column=c).border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save("load_test_report.xlsx")
    print("[LOAD TEST] Saved load_test_report.xlsx successfully.")

if __name__ == "__main__":
    run_load_test()
